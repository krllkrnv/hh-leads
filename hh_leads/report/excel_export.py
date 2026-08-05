"""Запись Excel-выгрузки из ChatRecord."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from hh_leads.classify import ACTION_PRIORITY, ChatRecord
from hh_leads.report.common import fmt_dt
from hh_leads.report.model import action_counts

COLUMNS = [
    "Компания",
    "Вакансия",
    "Ссылка на вакансию",
    "Ссылка на чат",
    "Статус",
    "Дата обновления",
    "Дата первого сообщения",
    "Краткое резюме переписки",
    "Признаки классификации",
    "Доп. категории",
]

ACTION_COLUMNS = [
    "Действие",
    "Детали",
    "Кто писал последним",
    "Компания",
    "Вакансия",
    "Статус",
    "Дата обновления",
    "Ссылка на чат",
    "Ссылка на вакансию",
    "Краткое резюме переписки",
]

ACTIONABLE = (
    "Ответить работодателю",
    "Собеседование / встреча",
    "Тестовое задание",
)


def record_row(rec: ChatRecord, category: str) -> list[Any]:
    return [
        rec.company,
        rec.vacancy_name,
        rec.vacancy_url,
        rec.chat_url,
        rec.state_name or rec.state_id,
        fmt_dt(rec.updated_at),
        fmt_dt(rec.first_message_at),
        rec.summary,
        rec.reasons_for(category),
        rec.extra_categories(category),
    ]


def action_row(rec: ChatRecord) -> list[Any]:
    return [
        rec.action,
        rec.action_detail,
        rec.last_from,
        rec.company,
        rec.vacancy_name,
        rec.state_name or rec.state_id,
        fmt_dt(rec.updated_at),
        rec.chat_url,
        rec.vacancy_url,
        rec.summary,
    ]


def autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[letter].width = max_len + 2


def _write_action_sheet(wb, title: str, records: list[ChatRecord], header_font: Font) -> None:
    ws = wb.create_sheet(title)
    for col, name in enumerate(ACTION_COLUMNS, start=1):
        cell = ws.cell(1, col, name)
        cell.font = header_font
    sorted_recs = sorted(
        records,
        key=lambda r: (
            ACTION_PRIORITY.get(r.action, 99),
            -(r.updated_at.timestamp() if r.updated_at else 0),
        ),
    )
    for i, rec in enumerate(sorted_recs, start=2):
        for col, value in enumerate(action_row(rec), start=1):
            cell = ws.cell(i, col, value)
            if col in (2, 8, 9, 10):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)
    ws.freeze_panes = "A2"


def write_excel(
    records: list[ChatRecord],
    out_path: Path,
    days: int,
    since: datetime,
) -> None:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    header_font = Font(bold=True)

    counts = {
        "Приглашения": sum(1 for r in records if "Приглашения" in r.categories),
        "Тестовые": sum(1 for r in records if "Тестовые" in r.categories),
        "Обсуждения": sum(1 for r in records if "Обсуждения" in r.categories),
    }
    action_counts_map = action_counts(records)
    multi = sum(1 for r in records if len(r.categories) > 1)

    summary_rows: list[tuple[str, Any]] = [
        ("Период", f"последние {days} дней (с {since.date().isoformat()})"),
        ("Дата выгрузки", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")),
        ("Всего чатов", len(records)),
        ("Приглашения", counts["Приглашения"]),
        ("Тестовые", counts["Тестовые"]),
        ("Обсуждения", counts["Обсуждения"]),
        ("В нескольких категориях", multi),
        ("", ""),
        ("—— Действия ——", ""),
    ]
    for action_name in sorted(action_counts_map.keys(), key=lambda a: ACTION_PRIORITY.get(a, 99)):
        summary_rows.append((action_name, action_counts_map[action_name]))

    ws_sum["A1"] = "Параметр"
    ws_sum["B1"] = "Значение"
    ws_sum["A1"].font = header_font
    ws_sum["B1"].font = header_font
    for i, (k, v) in enumerate(summary_rows, start=2):
        ws_sum[f"A{i}"] = k
        ws_sum[f"B{i}"] = v
    autosize(ws_sum)

    actionable = [r for r in records if r.action in ACTIONABLE]
    _write_action_sheet(wb, "Действия", actionable, header_font)

    def write_category_sheet(title: str) -> None:
        ws = wb.create_sheet(title)
        for col, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(1, col, name)
            cell.font = header_font
        rows = [r for r in records if title in r.categories]
        rows.sort(
            key=lambda r: r.updated_at or datetime.min.replace(tzinfo=timezone.utc),
            reverse=True,
        )
        for i, rec in enumerate(rows, start=2):
            for col, value in enumerate(record_row(rec, title), start=1):
                cell = ws.cell(i, col, value)
                if col in (3, 4, 8):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        autosize(ws)
        ws.freeze_panes = "A2"

    write_category_sheet("Приглашения")
    write_category_sheet("Тестовые")
    write_category_sheet("Обсуждения")
    _write_action_sheet(wb, "Все действия", records, header_font)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
