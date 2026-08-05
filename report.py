"""Отчёт для UI/Excel: сборка JSON-модели, запись и чтение xlsx."""

from __future__ import annotations

import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from classify import ACTION_PRIORITY, ChatRecord

COLUMNS = [
    "Компания",
    "Вакансия",
    "Ссылка на вакансию",
    "Ссылка на чат",
    "Статус",
    "Дата обновления",
    "Дата первого сообщения",
    "Краткое резюме переписки",
    "Признаки классификации",
    "Доп. категории",
]

ACTION_COLUMNS = [
    "Действие",
    "Детали",
    "Кто писал последним",
    "Компания",
    "Вакансия",
    "Статус",
    "Дата обновления",
    "Ссылка на чат",
    "Ссылка на вакансию",
    "Краткое резюме переписки",
]

STATE_ID_BY_NAME = {
    "Отклик": "response",
    "Приглашение": "invitation",
    "Собеседование": "interview",
    "Отказ": "discard",
    "Выход на работу": "hired",
    "Скрытый": "hidden",
}


def fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.astimezone().strftime("%Y-%m-%d %H:%M")


def parse_excel_dt(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def chat_id_from_url(url: str) -> str:
    m = re.search(r"/chat/(\d+)", str(url or ""))
    return m.group(1) if m else ""


def lead_tag(rec: ChatRecord) -> str:
    """Выбирает одну плашку для UI.

    Сначала отказ, потом текущее ожидание или автоответ, затем «позвоните»,
    «нужен ответ», собес, тест и приглашение из истории.
    """
    if rec.vacancy_closed or rec.action == "Отказ / закрыто":
        return "closed"
    # Уже ответили — важнее старых категорий «приглашение» или «тестовое».
    if rec.action == "Ждать ответа HR":
        return "wait"
    if rec.action == "Автоответ / бот":
        return "bot"
    # Явная просьба позвонить или написать важнее простого «ответьте на сообщение».
    if rec.strong_contact:
        return "call"
    if rec.action == "Ответить работодателю":
        return "reply"
    if (
        rec.state_id.lower() == "interview"
        or rec.state_name == "Собеседование"
        or rec.action == "Собеседование / встреча"
    ):
        return "interview"
    if "Тестовые" in rec.categories or rec.action == "Тестовое задание":
        return "test"
    if "Приглашения" in rec.categories:
        return "invite"
    return "discuss"


def clean_why(text: str) -> str:
    """Убирает служебные префиксы классификатора из колонки «суть»."""
    why = re.sub(r"\s+", " ", text).strip()
    why = re.sub(r"^(?:последнее от HR|бот|последнее от тебя)\s*[:·]\s*", "", why, flags=re.I)
    return why[:160].rstrip(" ,.;:") or "—"


def record_to_lead(rec: ChatRecord) -> dict[str, Any]:
    why = rec.action_detail or "; ".join(rec.invite_reasons[:2] or rec.test_reasons[:2]) or rec.summary
    why = clean_why(why or "")
    return {
        "id": rec.negotiation_id or chat_id_from_url(rec.chat_url),
        "company": rec.company or "—",
        "vacancy": rec.vacancy_name or "—",
        "status": rec.state_name or rec.state_id or "—",
        "stateId": rec.state_id,
        "tag": lead_tag(rec),
        "action": rec.action,
        "why": why,
        "summary": rec.summary,
        "updated": fmt_dt(rec.updated_at),
        "chatUrl": rec.chat_url,
        "vacancyUrl": rec.vacancy_url,
        "closed": bool(rec.vacancy_closed or rec.action == "Отказ / закрыто"),
        "strong": bool(rec.strong_contact),
        "categories": list(rec.categories),
        "inviteReasons": list(rec.invite_reasons),
        "testReasons": list(rec.test_reasons),
        "lastFrom": rec.last_from,
    }


def build_report(
    records: list[ChatRecord],
    *,
    days: int,
    since: datetime | None = None,
    source: str = "sync",
    incomplete: bool = False,
) -> dict[str, Any]:
    counts = {
        "invites": sum(1 for r in records if "Приглашения" in r.categories),
        "tests": sum(1 for r in records if "Тестовые" in r.categories),
        "discussions": sum(1 for r in records if "Обсуждения" in r.categories),
    }
    action_counts: dict[str, int] = {}
    for r in records:
        action_counts[r.action] = action_counts.get(r.action, 0) + 1

    state_counts: dict[str, int] = {}
    for r in records:
        key = r.state_name or r.state_id or "—"
        state_counts[key] = state_counts.get(key, 0) + 1

    leads = [record_to_lead(r) for r in records]
    reply = [l for l in leads if l["action"] == "Ответить работодателю"]
    interview_hh = [l for l in leads if l["tag"] == "interview" or l["status"] == "Собеседование"]
    contact = [l for l in leads if l["strong"]]
    tests = [l for l in leads if l["tag"] == "test" or "Тестовые" in (l.get("categories") or [])]
    invites = [l for l in leads if l["tag"] == "invite" or "Приглашения" in (l.get("categories") or [])]
    wait = [l for l in leads if l["tag"] == "wait"]
    bot = [l for l in leads if l["tag"] == "bot"]
    closed = [l for l in leads if l["tag"] == "closed" or l["closed"]]

    order = {
        "call": 0,
        "reply": 1,
        "interview": 2,
        "test": 3,
        "invite": 4,
        "wait": 5,
        "bot": 6,
        "discuss": 7,
        "closed": 8,
    }
    all_leads = sorted(
        leads,
        key=lambda x: (order.get(str(x.get("tag")), 9), x.get("updated") or ""),
    )

    period = f"Активность в чатах за последние {days} дн."
    period_from = None
    if since is not None:
        period_from = since.date().isoformat()
        period = f"Активность в чатах с {period_from}"
    if incomplete:
        period = f"{period} (загрузку остановили раньше времени)"

    exported_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")

    return {
        "meta": {
            "period": period,
            "periodFrom": period_from,
            "days": days,
            "exportedAt": exported_at,
            "source": source,
            "incomplete": incomplete,
            "total": len(records),
            "invites": counts["invites"],
            "tests": counts["tests"],
            "discussions": counts["discussions"],
            "multiCategory": sum(1 for r in records if len(r.categories) > 1),
            "actions": {
                "reply": action_counts.get("Ответить работодателю", 0),
                "interview": action_counts.get("Собеседование / встреча", 0),
                "test": action_counts.get("Тестовое задание", 0),
                "wait": action_counts.get("Ждать ответа HR", 0),
                "bot": action_counts.get("Автоответ / бот", 0),
                "closed": action_counts.get("Отказ / закрыто", 0),
            },
            "hhStatus": {
                "response": state_counts.get("Отклик", 0),
                "reject": state_counts.get("Отказ", 0),
                "interview": state_counts.get("Собеседование", 0),
            },
            "actionCounts": action_counts,
            "stateCounts": state_counts,
        },
        "leads": {
            "all": all_leads,
            "reply": reply,
            "interview": interview_hh,
            "contact": contact,
            "tests": tests,
            "invites": invites,
            "wait": wait,
            "bot": bot,
            "closed": closed,
        },
    }


def record_row(rec: ChatRecord, category: str) -> list[Any]:
    return [
        rec.company,
        rec.vacancy_name,
        rec.vacancy_url,
        rec.chat_url,
        rec.state_name or rec.state_id,
        fmt_dt(rec.updated_at),
        fmt_dt(rec.first_message_at),
        rec.summary,
        rec.reasons_for(category),
        rec.extra_categories(category),
    ]


def autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 60))
        ws.column_dimensions[letter].width = max_len + 2


def write_excel(
    records: list[ChatRecord],
    out_path: Path,
    days: int,
    since: datetime,
) -> None:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    header_font = Font(bold=True)

    counts = {
        "Приглашения": sum(1 for r in records if "Приглашения" in r.categories),
        "Тестовые": sum(1 for r in records if "Тестовые" in r.categories),
        "Обсуждения": sum(1 for r in records if "Обсуждения" in r.categories),
    }
    action_counts: dict[str, int] = {}
    for r in records:
        action_counts[r.action] = action_counts.get(r.action, 0) + 1
    multi = sum(1 for r in records if len(r.categories) > 1)

    summary_rows: list[tuple[str, Any]] = [
        ("Период", f"последние {days} дней (с {since.date().isoformat()})"),
        ("Дата выгрузки", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")),
        ("Всего чатов", len(records)),
        ("Приглашения", counts["Приглашения"]),
        ("Тестовые", counts["Тестовые"]),
        ("Обсуждения", counts["Обсуждения"]),
        ("В нескольких категориях", multi),
        ("", ""),
        ("—— Действия ——", ""),
    ]
    for action_name in sorted(action_counts.keys(), key=lambda a: ACTION_PRIORITY.get(a, 99)):
        summary_rows.append((action_name, action_counts[action_name]))

    ws_sum["A1"] = "Параметр"
    ws_sum["B1"] = "Значение"
    ws_sum["A1"].font = header_font
    ws_sum["B1"].font = header_font
    for i, (k, v) in enumerate(summary_rows, start=2):
        ws_sum[f"A{i}"] = k
        ws_sum[f"B{i}"] = v
    autosize(ws_sum)

    ws_act = wb.create_sheet("Действия", 1)
    for col, name in enumerate(ACTION_COLUMNS, start=1):
        cell = ws_act.cell(1, col, name)
        cell.font = header_font
    actionable = [
        r
        for r in records
        if r.action
        in (
            "Ответить работодателю",
            "Собеседование / встреча",
            "Тестовое задание",
        )
    ]
    actionable.sort(
        key=lambda r: (
            ACTION_PRIORITY.get(r.action, 99),
            -(r.updated_at.timestamp() if r.updated_at else 0),
        )
    )
    for i, rec in enumerate(actionable, start=2):
        row = [
            rec.action,
            rec.action_detail,
            rec.last_from,
            rec.company,
            rec.vacancy_name,
            rec.state_name or rec.state_id,
            fmt_dt(rec.updated_at),
            rec.chat_url,
            rec.vacancy_url,
            rec.summary,
        ]
        for col, value in enumerate(row, start=1):
            cell = ws_act.cell(i, col, value)
            if col in (2, 8, 9, 10):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws_act)
    ws_act.freeze_panes = "A2"

    def write_category_sheet(title: str) -> None:
        ws = wb.create_sheet(title)
        for col, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(1, col, name)
            cell.font = header_font
        rows = [r for r in records if title in r.categories]
        rows.sort(
            key=lambda r: r.updated_at or datetime.min.replace(tzinfo=timezone.utc),
            reverse=True,
        )
        for i, rec in enumerate(rows, start=2):
            for col, value in enumerate(record_row(rec, title), start=1):
                cell = ws.cell(i, col, value)
                if col in (3, 4, 8):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        autosize(ws)
        ws.freeze_panes = "A2"

    write_category_sheet("Приглашения")
    write_category_sheet("Тестовые")
    write_category_sheet("Обсуждения")

    ws_all = wb.create_sheet("Все действия")
    for col, name in enumerate(ACTION_COLUMNS, start=1):
        cell = ws_all.cell(1, col, name)
        cell.font = header_font
    all_sorted = sorted(
        records,
        key=lambda r: (
            ACTION_PRIORITY.get(r.action, 99),
            -(r.updated_at.timestamp() if r.updated_at else 0),
        ),
    )
    for i, rec in enumerate(all_sorted, start=2):
        row = [
            rec.action,
            rec.action_detail,
            rec.last_from,
            rec.company,
            rec.vacancy_name,
            rec.state_name or rec.state_id,
            fmt_dt(rec.updated_at),
            rec.chat_url,
            rec.vacancy_url,
            rec.summary,
        ]
        for col, value in enumerate(row, start=1):
            cell = ws_all.cell(i, col, value)
            if col in (2, 8, 9, 10):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws_all)
    ws_all.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def records_from_report(report: dict[str, Any]) -> tuple[list[ChatRecord], int, datetime]:
    """Собирает ChatRecord из JSON-отчёта дашборда для выгрузки в Excel."""
    meta = report.get("meta") if isinstance(report.get("meta"), dict) else {}
    try:
        days = max(1, min(180, int(meta.get("days") or 60)))
    except (TypeError, ValueError):
        days = 60

    period_from = meta.get("periodFrom")
    since: datetime | None = None
    if period_from:
        try:
            since = datetime.fromisoformat(str(period_from))
            if since.tzinfo is None:
                since = since.replace(tzinfo=timezone.utc)
        except ValueError:
            since = None
    if since is None:
        since = datetime.now(timezone.utc) - timedelta(days=days)

    leads = report.get("leads") if isinstance(report.get("leads"), dict) else {}
    items = leads.get("all") if isinstance(leads.get("all"), list) else []
    records: list[ChatRecord] = []
    for lead in items:
        if not isinstance(lead, dict):
            continue
        chat_url = str(lead.get("chatUrl") or "")
        cid = str(lead.get("id") or "") or chat_id_from_url(chat_url)
        status_name = str(lead.get("status") or "")
        state_id = str(lead.get("stateId") or STATE_ID_BY_NAME.get(status_name, "") or "")
        categories = lead.get("categories")
        if not isinstance(categories, list) or not categories:
            categories = ["Обсуждения"]
        invite_reasons = lead.get("inviteReasons")
        if not isinstance(invite_reasons, list):
            invite_reasons = []
        test_reasons = lead.get("testReasons")
        if not isinstance(test_reasons, list):
            test_reasons = []
        records.append(
            ChatRecord(
                negotiation_id=cid,
                company=str(lead.get("company") or ""),
                vacancy_name=str(lead.get("vacancy") or ""),
                vacancy_url=str(lead.get("vacancyUrl") or ""),
                chat_url=chat_url,
                state_id=state_id,
                state_name=status_name,
                updated_at=parse_excel_dt(lead.get("updated")),
                first_message_at=None,
                summary=str(lead.get("summary") or lead.get("why") or ""),
                invite_reasons=[str(x) for x in invite_reasons],
                test_reasons=[str(x) for x in test_reasons],
                categories=[str(c) for c in categories],
                action=str(lead.get("action") or "Без действия"),
                action_detail=str(lead.get("why") or ""),
                last_from=str(lead.get("lastFrom") or ""),
                strong_contact=bool(lead.get("strong")),
                vacancy_closed=bool(lead.get("closed")),
            )
        )
    return records, days, since


def _sheet_dicts(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
    return out


def records_from_excel(path: Path) -> tuple[list[ChatRecord], int]:
    """Восстанавливает ChatRecord из выгрузки скрипта. days — из сводки или 60."""
    wb = load_workbook(path, read_only=True, data_only=True)
    days = 60
    if "Сводка" in wb.sheetnames:
        for row in wb["Сводка"].iter_rows(values_only=True):
            if row and str(row[0] or "").startswith("Период"):
                m = re.search(r"(\d+)\s*дн", str(row[1] or ""))
                if m:
                    days = int(m.group(1))
                break

    by_id: dict[str, ChatRecord] = {}

    def upsert_from_action(row: dict[str, Any]) -> None:
        chat_url = str(row.get("Ссылка на чат") or "")
        cid = chat_id_from_url(chat_url)
        if not cid:
            return
        status_name = str(row.get("Статус") or "")
        state_id = STATE_ID_BY_NAME.get(status_name, status_name.lower())
        action = str(row.get("Действие") or "Без действия")
        detail = str(row.get("Детали") or "")
        summary = str(row.get("Краткое резюме переписки") or "")
        categories: list[str] = []
        invite_reasons: list[str] = []
        test_reasons: list[str] = []
        if action == "Собеседование / встреча":
            categories.append("Приглашения")
            invite_reasons.append(detail or "из листа Действия")
        if action == "Тестовое задание":
            categories.append("Тестовые")
            test_reasons.append(detail or "из листа Действия")
        if not categories:
            categories.append("Обсуждения")

        prev = by_id.get(cid)
        rec = ChatRecord(
            negotiation_id=cid,
            company=str(row.get("Компания") or (prev.company if prev else "")),
            vacancy_name=str(row.get("Вакансия") or (prev.vacancy_name if prev else "")),
            vacancy_url=str(row.get("Ссылка на вакансию") or (prev.vacancy_url if prev else "")),
            chat_url=chat_url or (prev.chat_url if prev else ""),
            state_id=state_id or (prev.state_id if prev else ""),
            state_name=status_name or (prev.state_name if prev else ""),
            updated_at=parse_excel_dt(row.get("Дата обновления"))
            or (prev.updated_at if prev else None),
            first_message_at=prev.first_message_at if prev else None,
            summary=summary or (prev.summary if prev else ""),
            invite_reasons=invite_reasons or (prev.invite_reasons if prev else []),
            test_reasons=test_reasons or (prev.test_reasons if prev else []),
            categories=list(dict.fromkeys((prev.categories if prev else []) + categories)),
            action=action,
            action_detail=detail,
            last_from=str(row.get("Кто писал последним") or (prev.last_from if prev else "")),
            strong_contact=bool(
                re.search(
                    r"позвон|напиш|свяжитесь|telegram|телеграм|whats|ватсап|созвон|\+7",
                    f"{detail} {summary}",
                    re.I,
                )
            ),
            vacancy_closed=bool(
                re.search(r"уже закрыл|вакансия закрыт|не актуал", f"{detail} {summary}", re.I)
            ),
        )
        by_id[cid] = rec

    def upsert_from_category(row: dict[str, Any], category: str) -> None:
        chat_url = str(row.get("Ссылка на чат") or "")
        cid = chat_id_from_url(chat_url)
        if not cid:
            return
        status_name = str(row.get("Статус") or "")
        state_id = STATE_ID_BY_NAME.get(status_name, status_name.lower())
        reasons = str(row.get("Признаки классификации") or "")
        prev = by_id.get(cid)
        cats = list(prev.categories) if prev else []
        if category not in cats:
            cats.append(category)
        invite_reasons = list(prev.invite_reasons) if prev else []
        test_reasons = list(prev.test_reasons) if prev else []
        if category == "Приглашения" and reasons:
            invite_reasons = [reasons]
        if category == "Тестовые" and reasons:
            test_reasons = [reasons]
        action = prev.action if prev else "Без действия"
        if category == "Приглашения" and action == "Без действия":
            action = "Собеседование / встреча"
        if category == "Тестовые" and action not in (
            "Ответить работодателю",
            "Собеседование / встреча",
        ):
            action = "Тестовое задание"

        by_id[cid] = ChatRecord(
            negotiation_id=cid,
            company=str(row.get("Компания") or (prev.company if prev else "")),
            vacancy_name=str(row.get("Вакансия") or (prev.vacancy_name if prev else "")),
            vacancy_url=str(row.get("Ссылка на вакансию") or (prev.vacancy_url if prev else "")),
            chat_url=chat_url or (prev.chat_url if prev else ""),
            state_id=state_id or (prev.state_id if prev else ""),
            state_name=status_name or (prev.state_name if prev else ""),
            updated_at=parse_excel_dt(row.get("Дата обновления"))
            or (prev.updated_at if prev else None),
            first_message_at=parse_excel_dt(row.get("Дата первого сообщения"))
            or (prev.first_message_at if prev else None),
            summary=str(row.get("Краткое резюме переписки") or (prev.summary if prev else "")),
            invite_reasons=invite_reasons,
            test_reasons=test_reasons,
            categories=cats,
            action=action,
            action_detail=prev.action_detail if prev else reasons,
            last_from=prev.last_from if prev else "",
            strong_contact=bool(prev.strong_contact)
            if prev
            else bool(
                re.search(
                    r"позвон|напиш|свяжитесь|telegram|телеграм|whats|ватсап|созвон|\+7",
                    reasons,
                    re.I,
                )
            ),
            vacancy_closed=prev.vacancy_closed if prev else False,
        )

    if "Все действия" in wb.sheetnames:
        for row in _sheet_dicts(wb["Все действия"]):
            upsert_from_action(row)
    elif "Действия" in wb.sheetnames:
        for row in _sheet_dicts(wb["Действия"]):
            upsert_from_action(row)

    for sheet_name in ("Приглашения", "Тестовые", "Обсуждения"):
        if sheet_name in wb.sheetnames:
            for row in _sheet_dicts(wb[sheet_name]):
                upsert_from_category(row, sheet_name)

    wb.close()
    return list(by_id.values()), days
