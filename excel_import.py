"""Восстановление ChatRecord из Excel-выгрузки."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from classify import CLOSED_RE, STRONG_CONTACT_RE, ChatRecord, match_reasons
from report_common import STATE_ID_BY_NAME, chat_id_from_url, parse_excel_dt


def _sheet_dicts(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
    return out


def _matches(patterns: list[re.Pattern[str]], text: str) -> bool:
    return bool(match_reasons(text, patterns))


def records_from_excel(path: Path) -> tuple[list[ChatRecord], int]:
    """Восстанавливает ChatRecord из выгрузки. days — из сводки или 60."""
    wb = load_workbook(path, read_only=True, data_only=True)
    days = 60
    if "Сводка" in wb.sheetnames:
        for row in wb["Сводка"].iter_rows(values_only=True):
            if row and str(row[0] or "").startswith("Период"):
                m = re.search(r"(\d+)\s*дн", str(row[1] or ""))
                if m:
                    days = int(m.group(1))
                break

    by_id: dict[str, ChatRecord] = {}

    def upsert_from_action(row: dict[str, Any]) -> None:
        chat_url = str(row.get("Ссылка на чат") or "")
        cid = chat_id_from_url(chat_url)
        if not cid:
            return
        status_name = str(row.get("Статус") or "")
        state_id = STATE_ID_BY_NAME.get(status_name, status_name.lower())
        action = str(row.get("Действие") or "Без действия")
        detail = str(row.get("Детали") or "")
        summary = str(row.get("Краткое резюме переписки") or "")
        categories: list[str] = []
        invite_reasons: list[str] = []
        test_reasons: list[str] = []
        if action == "Собеседование / встреча":
            categories.append("Приглашения")
            invite_reasons.append(detail or "из листа Действия")
        if action == "Тестовое задание":
            categories.append("Тестовые")
            test_reasons.append(detail or "из листа Действия")
        if not categories:
            categories.append("Обсуждения")

        corpus = f"{detail} {summary}"
        prev = by_id.get(cid)
        rec = ChatRecord(
            negotiation_id=cid,
            company=str(row.get("Компания") or (prev.company if prev else "")),
            vacancy_name=str(row.get("Вакансия") or (prev.vacancy_name if prev else "")),
            vacancy_url=str(row.get("Ссылка на вакансию") or (prev.vacancy_url if prev else "")),
            chat_url=chat_url or (prev.chat_url if prev else ""),
            state_id=state_id or (prev.state_id if prev else ""),
            state_name=status_name or (prev.state_name if prev else ""),
            updated_at=parse_excel_dt(row.get("Дата обновления"))
            or (prev.updated_at if prev else None),
            first_message_at=prev.first_message_at if prev else None,
            summary=summary or (prev.summary if prev else ""),
            invite_reasons=invite_reasons or (prev.invite_reasons if prev else []),
            test_reasons=test_reasons or (prev.test_reasons if prev else []),
            categories=list(dict.fromkeys((prev.categories if prev else []) + categories)),
            action=action,
            action_detail=detail,
            last_from=str(row.get("Кто писал последним") or (prev.last_from if prev else "")),
            strong_contact=_matches(STRONG_CONTACT_RE, corpus),
            vacancy_closed=_matches(CLOSED_RE, corpus),
        )
        by_id[cid] = rec

    def upsert_from_category(row: dict[str, Any], category: str) -> None:
        chat_url = str(row.get("Ссылка на чат") or "")
        cid = chat_id_from_url(chat_url)
        if not cid:
            return
        status_name = str(row.get("Статус") or "")
        state_id = STATE_ID_BY_NAME.get(status_name, status_name.lower())
        reasons = str(row.get("Признаки классификации") or "")
        prev = by_id.get(cid)
        cats = list(prev.categories) if prev else []
        if category not in cats:
            cats.append(category)
        invite_reasons = list(prev.invite_reasons) if prev else []
        test_reasons = list(prev.test_reasons) if prev else []
        if category == "Приглашения" and reasons:
            invite_reasons = [reasons]
        if category == "Тестовые" and reasons:
            test_reasons = [reasons]
        action = prev.action if prev else "Без действия"
        if category == "Приглашения" and action == "Без действия":
            action = "Собеседование / встреча"
        if category == "Тестовые" and action not in (
            "Ответить работодателю",
            "Собеседование / встреча",
        ):
            action = "Тестовое задание"

        by_id[cid] = ChatRecord(
            negotiation_id=cid,
            company=str(row.get("Компания") or (prev.company if prev else "")),
            vacancy_name=str(row.get("Вакансия") or (prev.vacancy_name if prev else "")),
            vacancy_url=str(row.get("Ссылка на вакансию") or (prev.vacancy_url if prev else "")),
            chat_url=chat_url or (prev.chat_url if prev else ""),
            state_id=state_id or (prev.state_id if prev else ""),
            state_name=status_name or (prev.state_name if prev else ""),
            updated_at=parse_excel_dt(row.get("Дата обновления"))
            or (prev.updated_at if prev else None),
            first_message_at=parse_excel_dt(row.get("Дата первого сообщения"))
            or (prev.first_message_at if prev else None),
            summary=str(row.get("Краткое резюме переписки") or (prev.summary if prev else "")),
            invite_reasons=invite_reasons,
            test_reasons=test_reasons,
            categories=cats,
            action=action,
            action_detail=prev.action_detail if prev else reasons,
            last_from=prev.last_from if prev else "",
            strong_contact=bool(prev.strong_contact)
            if prev
            else _matches(STRONG_CONTACT_RE, reasons),
            vacancy_closed=prev.vacancy_closed if prev else False,
        )

    if "Все действия" in wb.sheetnames:
        for row in _sheet_dicts(wb["Все действия"]):
            upsert_from_action(row)
    elif "Действия" in wb.sheetnames:
        for row in _sheet_dicts(wb["Действия"]):
            upsert_from_action(row)

    for sheet_name in ("Приглашения", "Тестовые", "Обсуждения"):
        if sheet_name in wb.sheetnames:
            for row in _sheet_dicts(wb[sheet_name]):
                upsert_from_category(row, sheet_name)

    wb.close()
    return list(by_id.values()), days
